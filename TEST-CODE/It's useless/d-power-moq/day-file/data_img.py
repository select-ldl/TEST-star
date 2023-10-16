
import openpyxl
from transformers import BertTokenizerFast, BertForSequenceClassification, AdamW
import os
import torch


def files():
    # 读取XLSX文件
    wb = openpyxl.load_workbook('input.xlsx')
    sheet = wb.active
    # 创建一个新的工作表
    new_sheet = wb.create_sheet('Output')
    # 定义一个字典，用于将名称归类到行数
    name_to_rows = {}
    # 遍历输入工作表中的每一行
    for row in sheet.iter_rows(values_only=True):
        # 获取每个单元格的值
        name = row[0]
        # 如果名称不在字典中，则添加到字典中并初始化一个列表
        if name not in name_to_rows:
            name_to_rows[name] = []
            # 将该行添加到相应的名称列表中
        name_to_rows[name].append(row)
    # 遍历字典中的每个名称和行列表
    for name, rows in name_to_rows.items():
        # 在输出工作表中创建一个新的行
        new_row = new_sheet.create_row(new_sheet.max_row + 1)
        # 在新行中添加名称
        new_row.append(name)
        # 遍历每个行列表并写入输出工作表中
        for row in rows:
            new_row = new_sheet.create_row(new_sheet.max_row + 1)
            for cell in row:
                new_row.append(cell)
            # 保存新的工作表

    wb.save('output.xlsx')
def file_en():
    # 加载预训练模型和权重

    model_name = 'bert-base-uncased'

    model = BertForSequenceClassification.from_pretrained(model_name)

    tokenizer = BertTokenizerFast.from_pretrained(model_name)

    # 英文句子

    english_sentence = "I love London."

    # 将英文句子转换为模型输入的格式

    input_ids = tokenizer.encode(english_sentence, return_tensors='pt')

    # 进行翻译，得到法文句子的输出ID

    output_ids = tokenizer.decode(torch.argmax(model(input_ids).logits)[0])

    print(output_ids)
def file_xlsxs_d():

    # 打开 XLSX 文件
    workbook = openpyxl.load_workbook(init_fille()[0])

    # 选择要拆分数据的列
    column_number = 1  # 假设要拆分第2列的数据

    # 选择工作表
    worksheet = workbook.active
    # 读取指定列的数据
    data = []
    for row in worksheet.iter_rows(min_row=2, max_row=170, min_col=column_number, max_col=column_number):
        for cell in row:
            data.append(cell.value)
            # 将数据按 \n 拆分并打印每行的第一个数据
    split_data = [data[i:j] for i, j in zip(range(len(data)), range(1, len(data) + 1, 2))]
    for row in split_data:
        first_item = row[0]
        print(first_item)

def init_fille():
    """读文件"""
    xml_files = []
    for root, dirs, files in os.walk(os.getcwd()):
        for file in files:
            if os.path.splitext(file)[1] == '.xlsx':
                xml_file = os.path.join(os.getcwd(), file)  # 文件
                """处理数据"""
                xml_files.append(xml_file)
    return xml_files
def file_xlsx():
    import openpyxl
    import re
    # 打开 XLSX 文件
    workbook = openpyxl.load_workbook(init_fille()[0])
    # 选择要拆分数据的行
    # 选择工作表
    worksheet = workbook.active
    with open('%s1.txt'%init_fille()[0][:-5], 'w+', encoding='utf8') as f1:
        # f1.write("img")
        for i in range(1, 178):
            row_number = i  # 假设要拆分第1行的数据
            # 读取指定行的数据
            data = worksheet.cell(row=row_number, column=1).value
            # 将数据按 \n 拆分并打印
            split_data = data.split('\n')
            ak=""
            for mm in split_data:
                ak=ak+mm
                ak=ak+"\t"
            f1.write("Box gauge"+ak)
            f1.write("\n")
            print(i)
            # match_res_size = re.match(('尺寸：'), split_data[0])
        f1.close()

if __name__ == '__main__':
    file_xlsx()
