# import json
import os
import csv
import numpy
import pandas as pd
import numpy as np
import openpyxl
import sys
# from openpyxl import load_workbook
def init_fille():
    """读文件"""
    xml_files = []
    for root, dirs, files in os.walk(os.getcwd()):
        for file in files:
            if os.path.splitext(file)[1] == '.xlsx':
                xml_file = os.path.join(os.getcwd(), file)  # 文件
                """处理数据"""
                xml_files.append(xml_file)
    return xml_files



# def file_w():
#     with open('%s.txt'%init_fille()[0][:].split(".")[0][-6:], 'w+', encoding='utf8') as f1:
#         try:
#             # print(avc[data_if(8390)])
#             f1.write("\n")
#         except KeyError:
#             f1.write("\n")
#         print("url")
#         f1.close()

# print(init_fille()[0][:].split(".")[0][-6:])
#
def okpp(data):
    # star=star
    text = np.str_(data)
    if text=='nan':
        text= ""
    else:
       text=(str(data))
    data_text=text.split("\n")
    # if data_text!=[]:
    #     for i in data_text:
    #         star=star+i+","
    # return star
    return data_text
def if_for(data):
    pack=[]
    for i in range(len(data)):
        if data[i] not in pack:
            pack.append(data[i])
    return pack
def xxxx():
    workbook=openpyxl.load_workbook(init_fille()[0])
    sheet = workbook['Sheet1']  # 修改为你想要处理的表单名称
    # wb = openpyxl.load_workbook(init_fille()[0])
    data = ["规格", "材质", "尺寸", "颜色", "多规格价格"]
    data_kk=[]
    data_name=[]
    # mkkkie = wb.active
    # init_fille()[0][:].split(".")[0][-6:]
    with open('%s1.txt' %init_fille()[0][:].split(".")[0][-5:], 'w+', encoding='utf8') as f1:
        m1=""
        m2=""
        m3=""
        m4=""
        for i in range(2,len(pd.read_excel(init_fille()[0], usecols=[0]).values)+2):
            row1 = sheet.cell(row=i, column=1).value  # 修改为你想要处理的行号和列号
            # 将行数据按照 \n 进行分隔
            row2 = sheet.cell(row=i, column=2).value  # 修改为你想要处理的行号和列号
            row3 = sheet.cell(row=i, column=3).value  # 修改为你想要处理的行号和列号
            row4 = sheet.cell(row=i, column=4).value  # 修改为你想要处理的行号和列号
            row5 = sheet.cell(row=i, column=5).value  # 修改为你想要处理的行号和列号
            row6 = sheet.cell(row=i, column=6).value
            # print(row1)
            a1=okpp(row1)
            a2 = okpp(row2)
            a3 = okpp(row3)
            a4 = okpp(row4)
            a5 = okpp(row5)
            pcl=[]
            lens=0
            stars=""
            star1=""
            star2=""
            star3=""
            star4=""
            if a1!=['None']:
                pcl.append(a1)
                lens = lens + len(a1)
                for i in if_for(a1):
                    stars=stars + i+","
                    if i==a1[-1]:
                        star1 = star1 + i
                    else:
                        star1=star1+i+","
            if a2!=['None']:
                pcl.append(a2)
                lens = lens + len(a2)
                for i in if_for(a2):
                    stars = stars + i + ","
                    if i == a2[-1]:
                        star2 = star2 + i
                    else:
                        star2 = star2 + i + ","
            if a3!=['None']:
                pcl.append(a3)
                lens = lens + len(a3)
                for i in if_for(a3):
                    stars = stars + i + ","
                    if i == a3[-1]:
                        star3 = star3 + i
                    else:
                        star3 = star3 + i + ","
            if a4!=['None']:
                pcl.append(a4)
                lens = lens + len(a4)
                for i in if_for(a4):
                    stars = stars + i + ","
                    if i == a4[-1]:
                        star4 = star4 + i
                    else:
                        star4 = star4 + i + ","
            if a5!=['None']:
                lens=lens+len(a5)
                pcl.append(a5)
            # print(lens)
            # print(len(pcl))
            # print(lens%len(pcl))
            prices=""
            for ak in range(int(lens/len(pcl))):
                for mmm in pcl:
                    # print(pcl[ak])
                    if mmm==pcl[-1]:
                        # print(type(mmm[ak]))
                        skp=numpy.float_(mmm[ak])
                        row6=numpy.float_(row6)
                        # print(skp,type(skp))
                        prices = prices+ str(round((skp)*float(row6),2))
                    else:
                        # print(len(mmm),ak)
                        prices =prices+ mmm[ak]+","
                # prices = prices + "\n"
                print(prices)
                if int(ak)==int(lens/len(pcl))-1:
                    continue
                else:
                    prices = prices+"\n"
            # print(type(row6))
            # print(prices)
                # prices=i[-1]
            abcd={"规格":star1, "材质":star2, "尺寸":star3, "颜色":star4, "多规格价格":prices}
            m1=m1+star1+"\n"
            m2=m2+star2+"\n"
            m3=m3+star3+"\n"
            m4=m4+star4+"\n"

            data_kk.append([star1,star2,star3,star4,prices])
            data_name.append(abcd)
            try:
                f1.write(stars)
                f1.write("\n")
                f1.write(prices)
                f1.write("\n")
                f1.write("\n")
            except KeyError:
                f1.write("请换行\n")
                print()
            # swie.save(init_fille()[0])
        print("-------规格-------")
        print(m1)

        print("-------材质-------")
        print(m2)

        print("-------尺寸-------")
        print(m3)
        print("-------颜色-------")
        print(m4)
        # for i in data:
        #     print(i)
        # print(init_fille()[0][:].split(".")[0][-5:],)
        f1.close()
    with open('xlsxmmss.csv', 'w', encoding='utf8', newline='') as f:
        writer = csv.DictWriter(f, data)
        writer.writeheader()
        writer.writerows(data_name)
        f.close()
    # mkkkie.__save__()
    # dt = pd.DataFrame(data_kk, columns=data)
    # with pd.ExcelWriter(init_fille()[0],mode='a') as ff:
    #     dt.to_excel(ff,sheet_name="wa",index=False,header=False,startrow=0,startcol=0)
    # dt = pd.DataFrame(data_kk, columns=data)
    # dt.to_excel("xxxxx.xlsx", sheet_name="wase",index=0)  # xlsx ## init_fille()[0]

if __name__ == '__main__':
    xxxx()
    # if_for(["white","Curry","white","Curry","white","Curry"])
    # print(if_for(["white","Curry","white","Curry","white","Curry"]))
    # input("输入任意字符结束程序")
    # sys.exit()
    # print(len(a))

