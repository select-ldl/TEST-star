import itertools
import os
import csv
import numpy
import pandas as pd
import numpy as np
import openpyxl


def init_fille():
    """读文件"""
    xml_files = []
    for root, dirs, files in os.walk(os.getcwd()):
        for file in files:
            if os.path.splitext(file)[1] == '.xlsx':
                xml_file = os.path.join(os.getcwd(), file)  # 文件
                """处理数据"""
                xml_files.append(xml_file)
    return xml_files
def okpp(data):
    """将多规格价格按行分开"""
    # star=star
    text = np.str_(data)
    if text=='nan':
        text= ""
    else:
       text=(str(data))
    data_text=text.split("\n")
    # if data_text!=[]:
    #     for i in data_text:
    #         star=star+i+","
    # return star
    # print(data_text)
    return data_text
def pppp(am):
    """将数据按逗号隔开开，并转换成列表形式返回
    :param am:
    :return:
    """
    km=[]
    for i in am:
        kss = []
        for m in i.split(","):
            kss.append(m.strip())
        km.append(kss)
    return km
def if_for(data):
    """将列表数据累加"""
    pack=[]
    for i in range(len(data)):
        if data[i] not in pack:
            pack.append(data[i])
    return pack
def mm_iffor(data):
    """将字符串数据重新排列"""
    pack = []
    for i in data:
        for m in i.split(','):
            if m not in pack:
                pack.append(m)
        # print(i.split(','))
    return pack
def ccc_if(data):
    """将列表数据重新排列"""
    if ',' in data[0]:
        return mm_iffor(data)
    else:
        return if_for(data)
def kkkkkkkk():
    """处理数据"""
    workbook=openpyxl.load_workbook(init_fille()[0])
    sheet = workbook['Sheet1']  # 修改为你想要处理的表单名称
    data = ["规格", "材质", "尺寸", "颜色", "多规格价格"]
    data_name=[]
    m1,m2,m3,m4="","","",""
    for i in range(2, len(pd.read_excel(init_fille()[0], usecols=[0]).values) + 2):
        row1 = sheet.cell(row=i, column=1).value  # 修改为你想要处理的行号和列号
        # 将行数据按照 \n 进行分隔
        row2 = sheet.cell(row=i, column=2).value  # 修改为你想要处理的行号和列号
        row3 = sheet.cell(row=i, column=3).value  # 修改为你想要处理的行号和列号
        row4 = sheet.cell(row=i, column=4).value  # 修改为你想要处理的行号和列号
        row5 = sheet.cell(row=i, column=5).value  # 修改为你想要处理的行号和列号
        a6 = sheet.cell(row=i, column=6).value
        a1 = okpp(row1)
        a2 = okpp(row2)
        a3m = okpp(row3)
        a4m = okpp(row4)
        a5 = okpp(row5)
        star1 = ""
        star2 = ""
        star3 = ""
        star4 = ""
        pkdat=[]
        lnes=0
        if a1 != ['None']:
            pkdat.append(a1)
            lnes+=len(a1)
            for i in ccc_if(a1):
                star1 = star1 + i.strip() + ","
        if a2 != ['None']:
            lnes += len(a2)
            pkdat.append(a2)
            for i in ccc_if(a2):
                star2 = star2 + i.strip() + ","
        if a3m != ['None']:
            lnes += len(a3m)
            pkdat.append(a3m)
            for i in ccc_if(a3m):
                star3 = star3 + i.strip() + ","
        if a4m != ['None']:
            lnes+=len(a4m)
            pkdat.append(a4m)
            for i in ccc_if(a4m):
                star4 = star4 + i.strip() + ","
        if a5 != ['None']:
            lnes+=len(a5)
            pkdat.append(a5)
        jkks=[]
        #将值进行切割
        for mk in range(len(pkdat)):
            jkks.append(pppp(pkdat[mk]))
        prices=[]
        for ak in range(int(lnes / len(pkdat))):
            kpi=[]
            for mmm in jkks:
                # print(type(mmm[ak]))
                if type(mmm[ak])==str:
                    kpi.append(list(mmm[ak]))
                    print((mmm[ak]))
                else:
                    kpi.append(mmm[ak])
            prices.append(kpi)
        mkdata=[]
        for pel in prices:
            # print(type(pel))
            conbit=list(itertools.product(*pel))
            mkdata.append(conbit)
                # print(mmm[ak])
                # print(mmm[ak],jkks[-1][-1])
        jjjjjpppp=''
        for p1 in mkdata:
            for p2 in p1:
                poestr=''
                for p3 in p2:
                    if p3 == p2[-1]:
                        poestr += str(round((numpy.float_(p3)) * numpy.float_(a6), 2))
                    else:
                        poestr+=p3+','
                jjjjjpppp+=poestr+'\n'
        # print(mkdata)
        # print(jjjjjpppp)
        # time.sleep(5)
        # akpoe=one(a1, a2, pppp(a3m),pppp(a4m), a5, a6)
        abcd={"规格":star1[:-1], "材质":star2[:-1], "尺寸":star3[:-1], "颜色":star4[:-1], "多规格价格":jjjjjpppp[:-1]}
        print(abcd)
        data_name.append(abcd)
        m1 += star1[:-1] + "\n"
        m2 += star2[:-1] + "\n"
        m3 += star3[:-1] + "\n"
        m4 += star4[:-1] + "\n"
        # print(one(a1, a2, pppp(a3m),pppp(a4m), a5, a6))
    print("-------规格-------")
    print(m1)

    print("-------材质-------")
    print(m2)

    print("-------尺寸-------")
    print(m3)
    print("-------颜色-------")
    print(m4)
    #"""写入CSV文件"""
    with open('xlsxsize.csv', 'w', encoding='utf8', newline='') as f:
        writer = csv.DictWriter(f, data)
        writer.writeheader()
        writer.writerows(data_name)
        f.close()
if __name__ == '__main__':

    kkkkkkkk()